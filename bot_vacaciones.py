# ============================================================
#  TechFlow S.A. — Bot de Gestión de Vacaciones
#  TPI — Programación y Sistemas de Información
#
#  INSTRUCCIONES:
#  1. Descargá este archivo y TechFlow_Vacaciones.xlsx
#  2. Ponelos en la MISMA carpeta
#  3. En VSCode abrí esa carpeta (File > Open Folder)
#  4. Abrí la terminal (Ctrl + J)
#  5. Escribí: python bot_vacaciones.py
# ============================================================

import openpyxl
from datetime import datetime, date
import os
import sys

# ============================================================
#  BUSCAR EL EXCEL AUTOMÁTICAMENTE
#  Busca en la carpeta del script y en la carpeta actual
# ============================================================

NOMBRE_EXCEL = "TechFlow_Vacaciones.xlsx"

def encontrar_excel():
    """
    Busca el Excel en:
    1. La carpeta donde está este script
    2. La carpeta desde donde se ejecuta Python
    """
    # Opción 1: misma carpeta que el script
    carpeta_script = os.path.dirname(os.path.abspath(__file__))
    ruta1 = os.path.join(carpeta_script, NOMBRE_EXCEL)
    if os.path.exists(ruta1):
        return ruta1

    # Opción 2: carpeta actual de trabajo
    ruta2 = os.path.join(os.getcwd(), NOMBRE_EXCEL)
    if os.path.exists(ruta2):
        return ruta2

    return None

EXCEL = encontrar_excel()

# ============================================================
#  COLORES CONSOLA
# ============================================================
VERDE    = "\033[92m"
ROJO     = "\033[91m"
AMARILLO = "\033[93m"
AZUL     = "\033[94m"
RESET    = "\033[0m"
NEGRITA  = "\033[1m"

# ============================================================
#  FUNCIONES QUE LEEN EL EXCEL
# ============================================================

def obtener_empleado(nombre):
    """
    Lee la hoja 'Empleados'.
    Columnas: ID | Nombre | Email | Área | Responsable | Días Disponibles | Días Tomados
    """
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Empleados"]
    for fila in ws.iter_rows(min_row=3, values_only=True):
        if fila[1] and str(fila[1]).strip().lower() == nombre.strip().lower():
            return {
                "id":               str(fila[0]),
                "nombre":           str(fila[1]),
                "email":            str(fila[2]),
                "area":             str(fila[3]),
                "responsable":      str(fila[4]),
                "dias_disponibles": int(fila[5]) if fila[5] else 0,
                "dias_tomados":     int(fila[6]) if fila[6] else 0,
            }
    return None


def hay_conflicto_equipo(fecha_ini, fecha_fin):
    """
    Compuerta XOR #2: hay conflicto si 3 o más compañeros
    tienen solicitudes APROBADAS que se solapan con el período.
    """
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Solicitudes"]
    ausentes = set()
    for fila in ws.iter_rows(min_row=3, values_only=True):
        if not fila[0]:
            continue
        if str(fila[8]).strip().upper() != "APROBADA":
            continue
        f_ini_sol = fila[4]
        f_fin_sol = fila[5]
        if isinstance(f_ini_sol, datetime):
            f_ini_sol = f_ini_sol.date()
        if isinstance(f_fin_sol, datetime):
            f_fin_sol = f_fin_sol.date()
        if isinstance(f_ini_sol, date) and isinstance(f_fin_sol, date):
            if not (fecha_fin < f_ini_sol or fecha_ini > f_fin_sol):
                ausentes.add(fila[1])
    return len(ausentes) >= 3


def proximo_numero():
    """Genera el próximo número de solicitud correlativo."""
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Solicitudes"]
    ultimo = 0
    for fila in ws.iter_rows(min_row=3, values_only=True):
        if fila[0] and str(fila[0]).startswith("VAC-"):
            try:
                num = int(str(fila[0]).split("-")[-1])
                if num > ultimo:
                    ultimo = num
            except:
                pass
    return f"VAC-2025-{ultimo + 1:04d}"


# ============================================================
#  FUNCIONES QUE ESCRIBEN EN EL EXCEL
# ============================================================

def registrar_solicitud(emp, fecha_ini, fecha_fin, dias, motivo, estado):
    """Agrega una fila en la hoja Solicitudes con formato y color."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    id_sol = proximo_numero()
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Solicitudes"]

    COLORES = {
        "APROBADA":  ("D1FAE5", "065F46"),
        "PENDIENTE": ("FEF3C7", "92400E"),
        "RECHAZADA": ("FEE2E2", "991B1B"),
    }
    bg_e, txt_e = COLORES.get(estado, ("FFFFFF", "1A1A2E"))

    borde = Border(
        left   = Side(style='thin', color="CCCCCC"),
        right  = Side(style='thin', color="CCCCCC"),
        top    = Side(style='thin', color="CCCCCC"),
        bottom = Side(style='thin', color="CCCCCC"),
    )

    fila_num = ws.max_row + 1
    datos = [id_sol, emp["id"], emp["nombre"], emp["area"],
             fecha_ini, fecha_fin, dias, motivo, estado,
             emp["responsable"], ""]
    ws.append(datos)

    for col_idx in range(1, len(datos) + 1):
        cell = ws.cell(row=fila_num, column=col_idx)
        cell.border    = borde
        cell.alignment = Alignment(
            horizontal = "center" if col_idx in [1, 2, 7] else "left",
            vertical   = "center"
        )
        cell.font = Font(name="Arial", size=10, color="1A1A2E")
        if col_idx in [5, 6]:
            cell.number_format = "DD/MM/YYYY"
        if col_idx == 9:
            cell.font = Font(name="Arial", size=10, bold=True, color=txt_e)
            cell.fill = PatternFill("solid", fgColor=bg_e)

    wb.save(EXCEL)
    return id_sol


def actualizar_saldo(nombre, dias_descontar):
    """Descuenta días disponibles y suma días tomados en la hoja Empleados."""
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Empleados"]
    for fila in ws.iter_rows(min_row=3):
        if fila[1].value and str(fila[1].value).strip().lower() == nombre.strip().lower():
            fila[5].value = int(fila[5].value or 0) - dias_descontar  # disponibles
            fila[6].value = int(fila[6].value or 0) + dias_descontar  # tomados
            break
    wb.save(EXCEL)


# ============================================================
#  UTILIDADES
# ============================================================

def parsear_fecha(texto):
    try:
        return datetime.strptime(texto.strip(), "%d/%m/%Y").date()
    except ValueError:
        return None

def dias_entre(a, b):
    return (b - a).days + 1

def sep():
    print(f"{AZUL}{'─' * 55}{RESET}")

def bot(msg):
    print(f"\n{VERDE}{NEGRITA}🤖 Bot:{RESET} {msg}")

def err(msg):
    print(f"\n{ROJO}⚠️  Error:{RESET} {msg}")

def ok(msg):
    print(f"\n{VERDE}✅{RESET} {msg}")

def warn(msg):
    print(f"\n{AMARILLO}⚠️{RESET} {msg}")

def pedir():
    return input("  Vos: ").strip()

def cancelado(t):
    return t.lower() == "/cancelar"


# ============================================================
#  FLUJO PRINCIPAL — MÁQUINA DE ESTADOS
# ============================================================

def iniciar_bot():
    print(f"\n{NEGRITA}{AZUL}")
    sep()
    print("  TECHFLOW S.A. — BOT DE GESTIÓN DE VACACIONES")
    print("  TPI · Programación y Sistemas de Información")
    sep()
    print(f"{RESET}")
    print("  Escribí /cancelar en cualquier momento para salir.")
    sep()

    # ── Verificar Excel ──────────────────────────────────────
    if not EXCEL:
        print(f"{ROJO}")
        print(f"  ⚠️  No se encontró: {NOMBRE_EXCEL}")
        print(f"  Asegurate de que esté en la misma carpeta que este script.")
        print(f"  Carpeta del script: {os.path.dirname(os.path.abspath(__file__))}")
        print(f"{RESET}")
        return

    print(f"  📂 Excel encontrado: {EXCEL}\n")

    # ── IDLE: pedir nombre ───────────────────────────────────
    bot("¡Hola! ¿Cuál es tu nombre completo?")
    nombre = pedir()
    if cancelado(nombre):
        warn("Solicitud cancelada.")
        return

    empleado = obtener_empleado(nombre)
    if not empleado:
        err(f"No encontré '{nombre}' en el Excel.")
        print("  Nombres disponibles:")
        print("   Maria Garcia · Carlos Lopez · Ana Torres · Pedro Romero")
        print("   Laura Diaz · Sofia Mendez · Diego Herrera · Valentina Cruz")
        return

    ok(f"Empleado encontrado: {empleado['nombre']}")
    print(f"  📊 Días disponibles: {empleado['dias_disponibles']}")
    print(f"  📊 Días tomados:     {empleado['dias_tomados']}")
    print(f"  👔 Responsable:      {empleado['responsable']}")

    # ── DATE_START ───────────────────────────────────────────
    while True:
        bot("¿Cuál es la fecha de INICIO? (DD/MM/AAAA)")
        entrada = pedir()
        if cancelado(entrada):
            warn("Solicitud cancelada. Datos no guardados.")
            return
        fecha_ini = parsear_fecha(entrada)
        if not fecha_ini:
            err("Formato inválido. Usá DD/MM/AAAA — Ej: 15/07/2025")
            continue
        if fecha_ini < date.today():
            err("La fecha de inicio no puede ser en el pasado.")
            continue
        break

    # ── DATE_END ─────────────────────────────────────────────
    while True:
        bot("¿Cuál es la fecha de FIN? (DD/MM/AAAA)")
        entrada = pedir()
        if cancelado(entrada):
            warn("Solicitud cancelada. Datos no guardados.")
            return
        fecha_fin = parsear_fecha(entrada)
        if not fecha_fin:
            err("Formato inválido. Usá DD/MM/AAAA — Ej: 25/07/2025")
            continue
        if fecha_fin <= fecha_ini:
            err("La fecha de fin debe ser posterior a la de inicio.")
            continue
        break

    dias = dias_entre(fecha_ini, fecha_fin)
    print(f"\n  📅 Período: {fecha_ini.strftime('%d/%m/%Y')} → {fecha_fin.strftime('%d/%m/%Y')}")
    print(f"  📊 Total:   {dias} días corridos")

    # ── REASON ───────────────────────────────────────────────
    bot("¿Cuál es el motivo? (o escribí OK para omitir)")
    motivo = pedir()
    if cancelado(motivo):
        warn("Solicitud cancelada. Datos no guardados.")
        return
    if not motivo or motivo.upper() == "OK":
        motivo = "Sin motivo especificado"

    # ── VALIDATING ───────────────────────────────────────────
    print(f"\n{AZUL}  ⚡ Validando en el Excel...{RESET}")

    # ── COMPUERTA XOR #1: ¿Tiene saldo? ─────────────────────
    print(f"\n  🔀 Compuerta XOR #1 — ¿Tiene saldo suficiente?")
    if empleado["dias_disponibles"] < dias:
        err(
            f"Saldo insuficiente.\n"
            f"  Días solicitados:  {dias}\n"
            f"  Días disponibles:  {empleado['dias_disponibles']}\n"
            f"  Podés solicitar hasta {empleado['dias_disponibles']} días."
        )
        registrar_solicitud(empleado, fecha_ini, fecha_fin, dias, motivo, "RECHAZADA")
        print(f"\n{ROJO}  📊 Registrado en Excel como RECHAZADA.{RESET}")
        return
    ok(f"Saldo OK ({empleado['dias_disponibles']} disponibles ≥ {dias} solicitados)")

    # ── COMPUERTA XOR #2: ¿Conflicto de equipo? ─────────────
    print(f"\n  🔀 Compuerta XOR #2 — ¿Hay conflicto de equipo?")
    if hay_conflicto_equipo(fecha_ini, fecha_fin):
        warn(
            "Alta carga de ausencias en ese período (3 o más compañeros).\n"
            "  Fechas alternativas sugeridas:\n"
            "   • Opción A: 01/08/2025 — 11/08/2025\n"
            "   • Opción B: 18/08/2025 — 28/08/2025\n"
            "  Iniciá una nueva solicitud con fechas alternativas."
        )
        return
    ok("Sin conflictos de equipo en ese período.")

    # ── PENDING ──────────────────────────────────────────────
    sep()
    print(f"\n  📤 Enviando solicitud a {empleado['responsable']}...")
    sep()
    print(f"   👤 Empleado:      {empleado['nombre']}")
    print(f"   📅 Inicio:        {fecha_ini.strftime('%d/%m/%Y')}")
    print(f"   📅 Fin:           {fecha_fin.strftime('%d/%m/%Y')}")
    print(f"   📊 Días:          {dias}")
    print(f"   💬 Motivo:        {motivo}")
    print(f"   👔 Responsable:   {empleado['responsable']}")
    sep()

    # ── COMPUERTA XOR #3: ¿El responsable aprueba? ──────────
    print(f"\n  🔀 Compuerta XOR #3 — Decisión del responsable del área")
    bot(f"¿El responsable ({empleado['responsable']}) aprueba? (SI / NO)")
    decision = pedir().upper()

    if decision == "SI":
        id_sol = registrar_solicitud(empleado, fecha_ini, fecha_fin, dias, motivo, "APROBADA")
        actualizar_saldo(empleado["nombre"], dias)
        sep()
        ok(f"¡SOLICITUD {id_sol} APROBADA y guardada en Excel!")
        print(f"\n  📋 Comprobante:")
        print(f"   Número:           {id_sol}")
        print(f"   Empleado:         {empleado['nombre']}")
        print(f"   Email:            {empleado['email']}")
        print(f"   Período:          {fecha_ini.strftime('%d/%m/%Y')} → {fecha_fin.strftime('%d/%m/%Y')}")
        print(f"   Días descontados: {dias}")
        print(f"   Saldo restante:   {empleado['dias_disponibles'] - dias} días")
        print(f"\n  ¡Que disfrutes tus vacaciones! 🌴")
    else:
        bot("¿Cuál es el motivo del rechazo?")
        motivo_rechazo = pedir() or "Sin motivo especificado"
        id_sol = registrar_solicitud(empleado, fecha_ini, fecha_fin, dias, motivo_rechazo, "RECHAZADA")
        sep()
        err(
            f"Solicitud {id_sol} RECHAZADA y registrada en Excel.\n"
            f"  Motivo: {motivo_rechazo}"
        )

    sep()
    print(f"\n  📂 Cambios guardados en: {EXCEL}\n")


# ============================================================
#  PUNTO DE ENTRADA
# ============================================================

if __name__ == "__main__":
    iniciar_bot()
